"""
Array Formula Puzzle Evaluator
LLM evaluation for Excel array formula puzzles

Evaluation Metrics:
- Exact Match: Exact match rate
- Numeric Accuracy: Accuracy within tolerance for numeric answers
- Type Accuracy: Accuracy by problem type

Data Format:
- Input: data/json/array_formula.jsonl
- Output: results/array_formula_results_{model}_{timestamp}.xlsx
"""

import json
import os
import re
import time
import argparse
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime
from dataclasses import dataclass, asdict

# API clients (optional imports)
try:
    import openai
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False

try:
    import anthropic
    ANTHROPIC_AVAILABLE = True
except ImportError:
    ANTHROPIC_AVAILABLE = False


@dataclass
class EvaluationResult:
    """Single problem evaluation result"""
    puzzle_id: str
    difficulty: str
    problem_type: str
    question: str
    expected_answer: Any
    model_answer: str
    parsed_answer: Any
    is_correct: bool
    is_numeric_close: bool
    error: Optional[str] = None
    latency_ms: float = 0.0


@dataclass
class EvaluationSummary:
    """Overall evaluation summary"""
    model: str
    total_problems: int
    correct_count: int
    accuracy: float
    numeric_close_accuracy: float
    by_difficulty: Dict[str, Dict[str, float]]
    by_type: Dict[str, Dict[str, float]]
    avg_latency_ms: float
    timestamp: str


# ============================================================
# Prompt Generation
# ============================================================

SYSTEM_PROMPT = """You are a spreadsheet/Excel expert.
Analyze the given table data and answer the question accurately.

Rules:
1. For numeric results, answer with only the number (no units, commas, or currency symbols)
2. For decimals, truncate unless otherwise specified
3. For text answers, provide the exact value only
4. Briefly explain your reasoning, then end with "Final answer: [answer]"
"""


def format_table_for_prompt(table_name: str, table_data: Dict) -> str:
    """Format table as markdown"""
    columns = table_data["columns"]
    data = table_data["data"]

    lines = [f"### {table_name} Table"]

    # Markdown table header
    header = "| " + " | ".join(str(col) for col in columns) + " |"
    separator = "|" + "|".join(["---"] * len(columns)) + "|"
    lines.append(header)
    lines.append(separator)

    # Data rows
    for row in data:
        row_str = "| " + " | ".join(str(row.get(col, "")) for col in columns) + " |"
        lines.append(row_str)

    return "\n".join(lines)


def puzzle_to_prompt(puzzle: Dict[str, Any], include_hint: bool = False) -> str:
    """Convert puzzle to LLM prompt"""
    prompt_parts = []

    prompt_parts.append("Analyze the following spreadsheet data.\n")

    # Table output
    for table_name, table_data in puzzle["tables"].items():
        prompt_parts.append(format_table_for_prompt(table_name, table_data))
        prompt_parts.append("")

    # Question
    prompt_parts.append(f"**Question**: {puzzle['question']}")

    # Hint (usually excluded during evaluation)
    if include_hint and "formula_hint" in puzzle:
        prompt_parts.append(f"**Hint**: {puzzle['formula_hint']}")

    # Response format instructions
    prompt_parts.append("\nBriefly explain your calculation, then end with 'Final answer: [answer]'.")
    if puzzle.get("answer_type") == "number":
        prompt_parts.append("(Answer with only a number, no units or commas)")

    return "\n".join(prompt_parts)


# ============================================================
# Answer Parsing
# ============================================================

def parse_answer(response: str, answer_type: str = "number") -> Any:
    """
    Extract final answer from LLM response

    Args:
        response: Full LLM response
        answer_type: "number" or "text"

    Returns:
        Parsed answer
    """
    patterns = [
        r"[Ff]inal\s*[Aa]nswer\s*[:：]\s*(.+?)(?:\n|$)",
        r"[Aa]nswer\s*[:：]\s*(.+?)(?:\n|$)",
    ]

    answer_text = None
    for pattern in patterns:
        match = re.search(pattern, response, re.IGNORECASE)
        if match:
            answer_text = match.group(1).strip()
            break

    # Fallback: extract from last line
    if answer_text is None:
        lines = [l.strip() for l in response.strip().split("\n") if l.strip()]
        if lines:
            answer_text = lines[-1]

    if answer_text is None:
        return None

    # Number type processing
    if answer_type == "number":
        number_match = re.search(r"-?[\d,]+\.?\d*", answer_text.replace(",", ""))
        if number_match:
            try:
                num_str = number_match.group().replace(",", "")
                if "." in num_str:
                    return float(num_str)
                return int(num_str)
            except ValueError:
                pass
        return None

    # Text type
    answer_text = answer_text.strip("'\"")
    return answer_text


def check_answer(
    expected: Any,
    parsed: Any,
    answer_type: str = "number",
    tolerance: float = 0.01
) -> Tuple[bool, bool]:
    """
    Check answer correctness

    Returns:
        (exact_match, numeric_close)
    """
    if parsed is None:
        return False, False

    if answer_type == "number":
        try:
            expected_num = float(expected)
            parsed_num = float(parsed)

            exact = abs(expected_num - parsed_num) < 0.001
            close = abs(expected_num - parsed_num) / max(abs(expected_num), 1) < tolerance

            return exact, close
        except (ValueError, TypeError):
            return False, False
    else:
        # Text comparison
        expected_str = str(expected).strip().lower()
        parsed_str = str(parsed).strip().lower()
        exact = expected_str == parsed_str
        return exact, exact


# ============================================================
# LLM API Call
# ============================================================

class LLMClient:
    """LLM API client wrapper"""

    def __init__(self, model: str, api_key: Optional[str] = None):
        self.model = model
        self.api_key = api_key

        if model.startswith("gpt") or model.startswith("o1"):
            if not OPENAI_AVAILABLE:
                raise ImportError("openai package required: pip install openai")
            self.client_type = "openai"
            self.client = openai.OpenAI(api_key=api_key or os.getenv("OPENAI_API_KEY"))
        elif model.startswith("claude"):
            if not ANTHROPIC_AVAILABLE:
                raise ImportError("anthropic package required: pip install anthropic")
            self.client_type = "anthropic"
            self.client = anthropic.Anthropic(api_key=api_key or os.getenv("ANTHROPIC_API_KEY"))
        else:
            raise ValueError(f"Unsupported model: {model}")

    def generate(self, prompt: str, system_prompt: str = SYSTEM_PROMPT) -> str:
        """Generate response for prompt"""
        if self.client_type == "openai":
            messages = [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": prompt}
            ]

            # o1 models don't support system messages
            if self.model.startswith("o1"):
                messages = [{"role": "user", "content": f"{system_prompt}\n\n{prompt}"}]

            response = self.client.chat.completions.create(
                model=self.model,
                messages=messages,
                temperature=0,
                max_tokens=2000
            )
            return response.choices[0].message.content

        elif self.client_type == "anthropic":
            response = self.client.messages.create(
                model=self.model,
                max_tokens=2000,
                system=system_prompt,
                messages=[{"role": "user", "content": prompt}]
            )
            return response.content[0].text

        return ""


# ============================================================
# Evaluation Execution
# ============================================================

def evaluate_puzzle(
    puzzle: Dict[str, Any],
    client: LLMClient,
    include_hint: bool = False
) -> EvaluationResult:
    """Evaluate single puzzle"""
    # New format: question field already contains prompt
    # Legacy format: need to generate prompt from tables
    if "question" in puzzle and "tables" not in puzzle:
        prompt = puzzle["question"]
    elif "tables" in puzzle:
        prompt = puzzle_to_prompt(puzzle, include_hint=include_hint)
    else:
        prompt = puzzle.get("question", "")

    if include_hint and puzzle.get("formula_hint"):
        prompt += f"\n\nHint: {puzzle['formula_hint']}"

    question_text = puzzle.get("question", prompt[:200])

    start_time = time.time()
    try:
        response = client.generate(prompt)
        latency_ms = (time.time() - start_time) * 1000

        parsed = parse_answer(response, puzzle.get("answer_type", "number"))
        exact, close = check_answer(
            puzzle["answer"],
            parsed,
            puzzle.get("answer_type", "number")
        )

        return EvaluationResult(
            puzzle_id=puzzle["id"],
            difficulty=puzzle["difficulty"],
            problem_type=puzzle.get("type", puzzle.get("problem_type", "unknown")),
            question=question_text,
            expected_answer=puzzle["answer"],
            model_answer=response,
            parsed_answer=parsed,
            is_correct=exact,
            is_numeric_close=close,
            latency_ms=latency_ms
        )

    except Exception as e:
        latency_ms = (time.time() - start_time) * 1000
        return EvaluationResult(
            puzzle_id=puzzle["id"],
            difficulty=puzzle["difficulty"],
            problem_type=puzzle.get("type", puzzle.get("problem_type", "unknown")),
            question=question_text,
            expected_answer=puzzle["answer"],
            model_answer="",
            parsed_answer=None,
            is_correct=False,
            is_numeric_close=False,
            error=str(e),
            latency_ms=latency_ms
        )


def evaluate_dataset(
    puzzles: List[Dict[str, Any]],
    client: LLMClient,
    include_hint: bool = False,
    verbose: bool = True,
    delay: float = 0.5
) -> Tuple[List[EvaluationResult], EvaluationSummary]:
    """Evaluate entire dataset"""
    results = []

    for i, puzzle in enumerate(puzzles):
        if verbose:
            print(f"[{i+1}/{len(puzzles)}] Evaluating {puzzle['id']}...", end=" ")

        result = evaluate_puzzle(puzzle, client, include_hint)
        results.append(result)

        if verbose:
            status = "O" if result.is_correct else ("~" if result.is_numeric_close else "X")
            print(f"{status} (expected: {result.expected_answer}, got: {result.parsed_answer})")

        if delay > 0 and i < len(puzzles) - 1:
            time.sleep(delay)

    summary = calculate_summary(results, client.model)

    return results, summary


def calculate_summary(results: List[EvaluationResult], model: str) -> EvaluationSummary:
    """Calculate evaluation summary"""
    total = len(results)
    correct = sum(1 for r in results if r.is_correct)
    numeric_close = sum(1 for r in results if r.is_numeric_close)
    avg_latency = sum(r.latency_ms for r in results) / total if total > 0 else 0

    # By difficulty
    by_difficulty = {}
    for diff in ["easy", "medium", "hard"]:
        diff_results = [r for r in results if r.difficulty == diff]
        if diff_results:
            by_difficulty[diff] = {
                "total": len(diff_results),
                "correct": sum(1 for r in diff_results if r.is_correct),
                "accuracy": sum(1 for r in diff_results if r.is_correct) / len(diff_results)
            }

    # By type
    by_type = {}
    types = set(r.problem_type for r in results)
    for ptype in types:
        type_results = [r for r in results if r.problem_type == ptype]
        if type_results:
            by_type[ptype] = {
                "total": len(type_results),
                "correct": sum(1 for r in type_results if r.is_correct),
                "accuracy": sum(1 for r in type_results if r.is_correct) / len(type_results)
            }

    return EvaluationSummary(
        model=model,
        total_problems=total,
        correct_count=correct,
        accuracy=correct / total if total > 0 else 0,
        numeric_close_accuracy=numeric_close / total if total > 0 else 0,
        by_difficulty=by_difficulty,
        by_type=by_type,
        avg_latency_ms=avg_latency,
        timestamp=datetime.now().isoformat()
    )


def print_summary(summary: EvaluationSummary):
    """Print evaluation summary"""
    print("\n" + "=" * 60)
    print(f"Evaluation Results - {summary.model}")
    print("=" * 60)
    print(f"Overall Accuracy: {summary.correct_count}/{summary.total_problems} ({summary.accuracy:.1%})")
    print(f"Numeric Close Accuracy: {summary.numeric_close_accuracy:.1%}")
    print(f"Average Latency: {summary.avg_latency_ms:.0f}ms")

    print("\nAccuracy by Difficulty:")
    for diff, stats in summary.by_difficulty.items():
        print(f"  {diff}: {stats['correct']}/{stats['total']} ({stats['accuracy']:.1%})")

    print("\nAccuracy by Type:")
    for ptype, stats in summary.by_type.items():
        print(f"  {ptype}: {stats['correct']}/{stats['total']} ({stats['accuracy']:.1%})")

    print("=" * 60)


def save_results_excel(
    results: List[EvaluationResult],
    summary: EvaluationSummary,
    output_dir: str = "../results"
):
    """
    Save evaluation results as Excel file

    Output path: results/array_formula_results_{model}_{timestamp}.xlsx
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("openpyxl package required: pip install openpyxl")
        # Fallback to JSON
        save_results_json(results, summary, output_dir)
        return

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    model_safe = summary.model.replace("/", "_").replace(":", "_")

    excel_file = output_path / f"array_formula_results_{model_safe}_{timestamp}.xlsx"

    wb = openpyxl.Workbook()

    # ============================================================
    # Sheet 1: Summary
    # ============================================================
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    correct_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    wrong_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Summary info
    summary_data = [
        ["Evaluation Summary", ""],
        ["Model", summary.model],
        ["Timestamp", summary.timestamp],
        ["Total Problems", summary.total_problems],
        ["Correct Count", summary.correct_count],
        ["Accuracy", f"{summary.accuracy:.1%}"],
        ["Numeric Close Accuracy", f"{summary.numeric_close_accuracy:.1%}"],
        ["Average Latency", f"{summary.avg_latency_ms:.0f}ms"],
        ["", ""],
        ["Accuracy by Difficulty", ""],
    ]

    for row_data in summary_data:
        ws_summary.append(row_data)

    # Difficulty stats
    ws_summary.append(["Difficulty", "Total", "Correct", "Accuracy"])
    for diff, stats in summary.by_difficulty.items():
        ws_summary.append([diff, stats['total'], stats['correct'], f"{stats['accuracy']:.1%}"])

    ws_summary.append(["", "", "", ""])
    ws_summary.append(["Accuracy by Type", "", "", ""])
    ws_summary.append(["Type", "Total", "Correct", "Accuracy"])
    for ptype, stats in summary.by_type.items():
        ws_summary.append([ptype, stats['total'], stats['correct'], f"{stats['accuracy']:.1%}"])

    # Column widths
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 20
    ws_summary.column_dimensions['C'].width = 10
    ws_summary.column_dimensions['D'].width = 15

    # ============================================================
    # Sheet 2: Details
    # ============================================================
    ws_details = wb.create_sheet("Details")

    # Header
    headers = ["ID", "Difficulty", "Type", "Question", "Expected", "Model Answer", "Parsed", "Correct", "Close", "Latency(ms)", "Error"]
    ws_details.append(headers)

    # Header style
    for col in range(1, len(headers) + 1):
        cell = ws_details.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Data rows
    for i, result in enumerate(results, 2):
        row = [
            result.puzzle_id,
            result.difficulty,
            result.problem_type,
            result.question[:100] + "..." if len(result.question) > 100 else result.question,
            str(result.expected_answer),
            result.model_answer[:200] + "..." if len(result.model_answer) > 200 else result.model_answer,
            str(result.parsed_answer) if result.parsed_answer is not None else "N/A",
            "O" if result.is_correct else "X",
            "O" if result.is_numeric_close else "X",
            f"{result.latency_ms:.0f}",
            result.error or ""
        ]
        ws_details.append(row)

        # Correct/wrong colors
        correct_cell = ws_details.cell(row=i, column=8)
        if result.is_correct:
            correct_cell.fill = correct_fill
        else:
            correct_cell.fill = wrong_fill

    # Column widths
    col_widths = [20, 10, 25, 50, 15, 50, 15, 10, 10, 12, 30]
    for col, width in enumerate(col_widths, 1):
        ws_details.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # Add filter
    ws_details.auto_filter.ref = ws_details.dimensions

    # ============================================================
    # Sheet 3: Full Responses
    # ============================================================
    ws_full = wb.create_sheet("Full Responses")

    ws_full.append(["ID", "Question", "Full Model Response"])
    for col in range(1, 4):
        cell = ws_full.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill

    for result in results:
        ws_full.append([
            result.puzzle_id,
            result.question,
            result.model_answer
        ])

    ws_full.column_dimensions['A'].width = 20
    ws_full.column_dimensions['B'].width = 80
    ws_full.column_dimensions['C'].width = 100

    # Save
    wb.save(excel_file)

    print(f"\nResults saved: {excel_file}")


def save_results_json(
    results: List[EvaluationResult],
    summary: EvaluationSummary,
    output_dir: str = "../results"
):
    """Save evaluation results as JSON (fallback)"""
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    model_safe = summary.model.replace("/", "_").replace(":", "_")

    # Save detailed results
    results_file = output_path / f"array_formula_results_{model_safe}_{timestamp}.json"
    with open(results_file, "w", encoding="utf-8") as f:
        json.dump([asdict(r) for r in results], f, ensure_ascii=False, indent=2)

    # Save summary
    summary_file = output_path / f"array_formula_summary_{model_safe}_{timestamp}.json"
    with open(summary_file, "w", encoding="utf-8") as f:
        json.dump(asdict(summary), f, ensure_ascii=False, indent=2)

    print(f"\nResults saved: {results_file}")
    print(f"Summary saved: {summary_file}")


# ============================================================
# JSONL Data Loading
# ============================================================

def load_puzzles_from_jsonl(jsonl_path: str) -> List[Dict[str, Any]]:
    """Load puzzle data from JSONL file"""
    puzzles = []
    with open(jsonl_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                puzzles.append(json.loads(line))
    return puzzles


# ============================================================
# CLI
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="Array Formula Puzzle Evaluator")
    parser.add_argument("--model", type=str, default="gpt-4o", help="Model to evaluate")
    parser.add_argument("--dataset", type=str, default="../data/json/array_formula.jsonl", help="Dataset path (JSONL)")
    parser.add_argument("--output", type=str, default="../results", help="Output directory")
    parser.add_argument("--limit", type=int, default=None, help="Limit number of puzzles")
    parser.add_argument("--difficulty", type=str, default=None, help="Filter by difficulty")
    parser.add_argument("--type", type=str, default=None, help="Filter by type")
    parser.add_argument("--hint", action="store_true", help="Include hints")
    parser.add_argument("--delay", type=float, default=0.5, help="Delay between API calls (seconds)")
    parser.add_argument("--quiet", action="store_true", help="Suppress verbose output")

    args = parser.parse_args()

    # Path setup
    script_dir = Path(__file__).parent
    dataset_path = script_dir / args.dataset

    # Load dataset
    print(f"Loading dataset: {dataset_path}")

    if not dataset_path.exists():
        print(f"Dataset file not found: {dataset_path}")
        print(f"Generate puzzles first:")
        print(f"  cd ../guess")
        print(f"  python array_formula.py")
        return

    # Load JSONL or JSON file
    if str(dataset_path).endswith(".jsonl"):
        puzzles = load_puzzles_from_jsonl(dataset_path)
    else:
        with open(dataset_path, "r", encoding="utf-8") as f:
            puzzles = json.load(f)

    # Filter
    if args.difficulty:
        puzzles = [p for p in puzzles if p["difficulty"] == args.difficulty]
    if args.type:
        puzzles = [p for p in puzzles if p["type"] == args.type]
    if args.limit:
        puzzles = puzzles[:args.limit]

    print(f"Puzzles to evaluate: {len(puzzles)}")

    if len(puzzles) == 0:
        print("No puzzles to evaluate.")
        return

    # Initialize client
    client = LLMClient(args.model)

    # Run evaluation
    results, summary = evaluate_dataset(
        puzzles,
        client,
        include_hint=args.hint,
        verbose=not args.quiet,
        delay=args.delay
    )

    # Output and save results
    print_summary(summary)

    output_dir = script_dir / args.output
    save_results_excel(results, summary, str(output_dir))


if __name__ == "__main__":
    main()
